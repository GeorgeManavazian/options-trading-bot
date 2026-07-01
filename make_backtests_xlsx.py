# make_backtests_xlsx.py — builds a polished, shareable Backtests.xlsx: a "Journey" story sheet
# + one detailed sheet per backtest, with conditional-formatting color. Re-run to regenerate /
# add a sheet when a new backtest lands.  Run:  .venv/bin/python make_backtests_xlsx.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

# ---- palette ----
NAVY = "1F3864"; BLUE = "2E5395"; LIGHT = "D9E1F2"; BAND = "F2F5FB"
GOOD_F, GOOD_T = "C6EFCE", "006100"
WARN_F, WARN_T = "FFEB9C", "9C6500"
BAD_F,  BAD_T  = "FFC7CE", "9C0006"
WHITE = "FFFFFF"; GREY = "808080"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_): return PatternFill("solid", fgColor=hex_)
def cell(ws, r, c, v, bold=False, size=11, color="000000", bg=None, align="left", border=True, wrap=False):
    x = ws.cell(r, c, v)
    x.font = Font(bold=bold, size=size, color=color, name="Calibri")
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg: x.fill = fill(bg)
    if border: x.border = BORDER
    return x


def verdict(ws, r, c, text, kind):
    f, t = {"good": (GOOD_F, GOOD_T), "warn": (WARN_F, WARN_T), "bad": (BAD_F, BAD_T)}[kind]
    return cell(ws, r, c, text, bold=True, color=t, bg=f, align="center")


def title(ws, r, text, sub=None, span=7):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    cell(ws, r, 1, text, bold=True, size=16, color=WHITE, bg=NAVY, align="left", border=False)
    ws.row_dimensions[r].height = 26
    if sub:
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=span)
        cell(ws, r + 1, 1, sub, size=10, color="595959", align="left", border=False)
        return r + 3
    return r + 2


wb = Workbook()

# ============================================================ Sheet 1: JOURNEY
ws = wb.active; ws.title = "Journey"
ws.sheet_view.showGridLines = False
row = title(ws, 1, "Options Trading Bot — The Backtest Journey",
            "Goal: automate a real trading edge. Philosophy: build it honestly, test it hard, and let "
            "the data — not the idea — decide. Most attempts SHOULD fail; proving that cheaply is the skill.", span=7)

heads = ["#", "Strategy", "What it is", "Backtest verdict", "Headline result", "Survives costs?", "Status"]
for c, h in enumerate(heads, 1):
    cell(ws, row, c, h, bold=True, color=WHITE, bg=BLUE, align="center")
ws.row_dimensions[row].height = 20
r = row + 1

journey = [
    ("1", "1DTE Iron Condor", "SPX 1-day condor (4 legs), sell premium",
     "LOST money", "77% win vs 79% breakeven · −396% on risk · 533% drawdown", "n/a", ("Shelved", "bad")),
    ("2", "ORB (Kirk let-expire)", "SPX 0DTE breakout → credit spread, let expire",
     "First real edge", "85% win vs 82% BE · +760% on risk", "YES", ("Shelved — savage drawdown, regime-dependent", "warn")),
    ("3", "ACD momentum (multi-day)", "Breakout continuation, directional",
     "No edge (fragment)", "Longs ≈ market beta · shorts negative", "n/a", ("Falsified", "bad")),
    ("4", "ACD 0DTE", "ACD entry → 0DTE credit spread",
     "Edge too thin", "86% win vs 85% BE · +205% @0¢ → dies at 5¢", "NO", ("Rejected — not cost-robust", "bad")),
    ("5", "★ FULL ACD Bot (multiday)", "Complete Fisher method → 30-DTE options, held 5 days",
     "REAL cost-robust edge", "+476% on risk · +428% even @20¢/leg", "YES", ("REFINE (55/100) — great edge, savage drawdown", "warn")),
]
for j in journey:
    num, strat, what, ver, res, cost, (stat, kind) = j
    star = j[1].startswith("★")
    cell(ws, r, 1, num, bold=True, align="center", bg=(LIGHT if star else None))
    cell(ws, r, 2, strat, bold=star, bg=(LIGHT if star else None))
    cell(ws, r, 3, what, size=10, wrap=True)
    vk = "good" if "REAL" in ver or "First" in ver else ("bad" if "LOST" in ver or "No edge" in ver or "thin" in ver else "warn")
    verdict(ws, r, 4, ver, vk)
    cell(ws, r, 5, res, size=10, wrap=True)
    ck = {"YES": "good", "NO": "bad", "n/a": None}[cost]
    if ck: verdict(ws, r, 6, cost, ck)
    else: cell(ws, r, 6, cost, align="center", color=GREY)
    verdict(ws, r, 7, stat, kind)
    ws.row_dimensions[r].height = 34
    r += 1

# takeaway banner
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
cell(ws, r, 1, "KEY LESSON, learned 3×: a real multi-day edge is destroyed by a too-tight exit. "
     "Hold to the signal's horizon. The FULL method — not any single fragment — is what finally worked.",
     bold=True, color=WARN_T, bg=WARN_F, align="left", wrap=True)
ws.row_dimensions[r].height = 30

widths = [4, 26, 34, 20, 40, 15, 34]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A5"

# ============================================================ Sheet 2: ACD FULL BOT
ws = wb.create_sheet("★ ACD Full Bot")
ws.sheet_view.showGridLines = False
row = title(ws, 1, "★ Full Mark Fisher ACD Bot — Backtest",
            "SPX · 2023-07 → 2026-06 (3 yrs) · multiday macro signals (sushi + reversal) filtered by the "
            "number-line chop filter & regime, expressed as ~30-DTE options, held to a 5-day horizon.", span=6)

# --- headline metrics grid ---
cell(ws, row, 1, "HEADLINE", bold=True, color=WHITE, bg=BLUE, align="center")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1
metrics = [("Trades", "102", "warn"), ("Win rate", "50%", "warn"),
           ("Total return (on risk)", "+476%", "good"), ("Max drawdown", "743% of risk cap", "bad"),
           ("Risk-adjusted (tot/DD)", "+0.64", "warn"), ("Sharpe-like", "+0.07", "bad"),
           ("Expectancy / trade", "+4.4%", "good"), ("Avg win / avg loss", "+52% / −44%", "warn"),
           ("Cost-robust?", "YES (to 20¢/leg)", "good"), ("backtest-expert grade", "55/100 — REFINE", "warn")]
mr = row
for i, (k, v, kind) in enumerate(metrics):
    rr = mr + (i // 2); cc = 1 + (i % 2) * 3
    cell(ws, rr, cc, k, bold=True, bg=BAND)
    ws.merge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc + 1)
    verdict(ws, rr, cc + 2, v, kind)
    ws.row_dimensions[rr].height = 19
row = mr + (len(metrics) + 1) // 2 + 1

def section(r, label):
    cell(ws, r, 1, label, bold=True, color=WHITE, bg=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 20
    return r + 1

# --- exit comparison ---
row = section(row, "① Exit choice = the whole ballgame (same signals, same options)")
for c, h in enumerate(["Exit rule", "Total on risk", "Win rate", "Read"], 1):
    cell(ws, row, c, h, bold=True, bg=LIGHT, align="center")
ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
row += 1
exits = [("3-day pivot TRAILING stop", "−163%", "35%", "LOSES — bails before the edge appears", "bad"),
         ("Fixed hold 5 days", "+476%", "50%", "WINS — held to the horizon", "good"),
         ("Fixed hold 10 days", "+297%", "39%", "Wins, but past the sweet spot", "warn")]
for name, tot, win, read, kind in exits:
    cell(ws, row, 1, name, bold=(kind == "good"))
    verdict(ws, row, 2, tot, kind); cell(ws, row, 3, win, align="center")
    cell(ws, row, 4, read, size=10); ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    row += 1

# --- slippage sweep ---
row += 1
row = section(row, "② Does the edge survive costs?  (YES — barely erodes)")
for c, h in enumerate(["Slippage / leg", "Total on risk", "Bar"], 1):
    cell(ws, row, c, h, bold=True, bg=LIGHT, align="center")
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
row += 1
slip_start = row
for lab, val in [("0¢", 476), ("5¢", 464), ("10¢", 452), ("20¢", 428)]:
    cell(ws, row, 1, lab, align="center")
    cell(ws, row, 2, f"+{val}%", bold=True, color=GOOD_T, bg=GOOD_F, align="center")
    cell(ws, row, 3, val, align="center"); ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    row += 1
ws.conditional_formatting.add(f"C{slip_start}:C{row-1}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=500, color="63BE7B"))

# --- per-setup ---
row += 1
row = section(row, "③ Where the edge lives (per setup type)")
for c, h in enumerate(["Setup", "Trades", "Win", "Total on risk"], 1):
    cell(ws, row, c, h, bold=True, bg=LIGHT, align="center")
row += 1
setups = [("sushi (outside-reversal)", "78", "54%", "+532%", "good"),
          ("reversal_trade (Fisher's best)", "23", "39%", "+14%", "warn"),
          ("trt", "1", "0%", "−70%", "bad")]
for name, n, win, tot, kind in setups:
    cell(ws, row, 1, name); cell(ws, row, 2, n, align="center"); cell(ws, row, 3, win, align="center")
    verdict(ws, row, 4, tot, kind); ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    row += 1

# --- red flags / verdict ---
row += 1
row = section(row, "④ Why 'REFINE', not 'Deploy' — the honest red flags")
flags = ["🚩 Max drawdown 743% of risk capital — catastrophic path; tradeable only at ~1% size per trade.",
         "🚩 8 tunable params + the 5-day hold was chosen from the checkpoint → overfitting / circularity risk (needs out-of-sample).",
         "🚩 Only 3 years tested — one regime-ish window; needs 5+ and walk-forward.",
         "✔ But: real positive expectancy, cost-robust, and it VINDICATES building the whole method — the fragment hid this edge."]
for f in flags:
    kind = GOOD_F if f.startswith("✔") else BAD_F
    tcol = GOOD_T if f.startswith("✔") else BAD_T
    cell(ws, row, 1, f, size=10, color=tcol, bg=kind, wrap=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 1

for c, w in enumerate([30, 16, 12, 16, 12, 12], 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A5"

wb.save("Backtests.xlsx")
print("wrote Backtests.xlsx  (sheets: Journey, ★ ACD Full Bot)")
